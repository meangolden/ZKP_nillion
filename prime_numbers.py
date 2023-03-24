# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 23:47:55 2023

@author: cp17593
"""
import openpyxl


# Generate safe primes from Sophie Germain primes (ref: https://oeis.org/A005384)
from domain_parameters import count_digits, find_primitives

sophie_germain_primes = [ 11, 23, 29, 41, 53, 83, 89, 113, 131,\
                         173, 179, 191, 233, 239, 251, 281, 293, 359, 419,\
                         431, 443, 491, 509, 593, 641, 653, 659, 683, 719,\
                         743, 761, 809, 911, 953, 1013, 1019, 1031, 1049,\
                         1103, 1223, 1229, 1289, 1409, 1439, 1451, 1481, \
                         1499, 1511, 1559]

no_digits = [count_digits(q)  for q in sophie_germain_primes] 

safe_primes = [2 * q + 1 for q in sophie_germain_primes]
  
# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Add first sheet and column headers
sheet1 = workbook.active
sheet1.title = "exp domain"
sheet1.cell(row=1, column=1, value="q-size")
sheet1.cell(row=1, column=2, value="q")
sheet1.cell(row=1, column=3, value="p")
sheet1.cell(row=1, column=4, value="g")
sheet1.cell(row=1, column=5, value="h")

# Initialize row counter
row_num = 2
  
for i in range(len(safe_primes)- 1): # it takes long for the last large prime
    sheet1.cell(row=row_num, column=1, value=no_digits[i])
    sheet1.cell(row=row_num, column=2, value=sophie_germain_primes[i])
    sheet1.cell(row=row_num, column=3, value=safe_primes[i])
    
    g, h = find_primitives(sophie_germain_primes[i], safe_primes[i])
    sheet1.cell(row=row_num, column=4, value=g)
    sheet1.cell(row=row_num, column=5, value=h)
      
    row_num += 1       

# Save the workbook
#workbook.save("domain_parameters.xlsx")
#os.startfile('domain_parameters.xlsx')

#large_q = 1846389521368 + 11**600
#large_p = large_q *2 + 1
